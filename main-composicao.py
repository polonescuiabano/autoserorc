import openpyxl
from pymongo import MongoClient
from datetime import datetime

# Conectar ao MongoDB
client = MongoClient('mongodb://localhost:27017/')  # Altere a URL de conexão conforme necessário
db = client['serorc']  # Substitua pelo nome do seu banco de dados
composicoes_collection = db['composicoes']


# Função para adicionar uma composição principal no MongoDB
def adicionar_composicao_principal(descricao_classe, sigla_classe, codigo, descricao, unidade_medida, tipo,
                                   data_cotacao, preco_desonerado):
    # Verificar se os dados essenciais estão presentes e válidos
    if not descricao_classe or not sigla_classe or not codigo or not descricao or not unidade_medida or preco_desonerado is None:
        print(
            f"Dados inválidos para inserção: {descricao_classe}, {sigla_classe}, {codigo}, {descricao}, {unidade_medida}, {preco_desonerado}")
        return None  # Não insere a composição se os dados estiverem faltando ou forem inválidos

    composicao = {
        "tipo": tipo,
        "descricao_classe": descricao_classe,
        "sigla_classe": sigla_classe,
        "codigo": codigo,
        "descricao": descricao,
        "unidade_medida": unidade_medida,
        "precos_cotacao": [
            {
                "preco_desonerado": preco_desonerado,  # Preço desonerado adicionado
                "preco_nao_desonerado": None,  # Preço não desonerado será adicionado depois
                "data_cotacao": data_cotacao
            }
        ],
        "composicoes_auxiliares": [],
        "insumos": []
    }

    # Adicionar a composição à coleção
    result = composicoes_collection.insert_one(composicao)
    return result.inserted_id


# Função para processar o preço desonerado
def processar_preco(preco_desonerado):
    preco_str = str(preco_desonerado).strip()

    # Remover ponto como separador de milhar
    preco_str = preco_str.replace(".", "")

    # Substituir a vírgula por ponto para garantir que seja tratado como decimal
    preco_str = preco_str.replace(",", ".")

    try:
        # Tentar converter o valor para float
        return float(preco_str)
    except ValueError:
        print(f"Valor inválido para conversão para float: {preco_str}. Ignorando.")
        return None


# Função para ler e processar o arquivo Excel
def processar_composicoes_excel(caminho_arquivo):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(caminho_arquivo)
    sheet = wb.active  # Pegando a primeira aba da planilha

    # Definir a data de cotação e o tipo
    data_cotacao = datetime(2024, 11, 1)
    tipo = "SINAPI"

    # Iterar pelas linhas do Excel a partir da linha 7 até a linha 7822
    for row in range(7, 10000):  # Ajustado para começar na linha 7 e terminar na linha 7822
        descricao_classe = sheet.cell(row=row, column=1).value
        sigla_classe = sheet.cell(row=row, column=2).value
        codigo = sheet.cell(row=row, column=7).value
        descricao = sheet.cell(row=row, column=8).value
        unidade_medida = sheet.cell(row=row, column=9).value
        preco_desonerado = sheet.cell(row=row, column=11).value  # Preço desonerado da coluna K

        # Verificar se o código não está vazio e convertê-lo para inteiro
        if codigo is not None:
            codigo = int(codigo)

        # Processar o preço desonerado
        preco_desonerado = processar_preco(preco_desonerado)

        # Verificar se os dados essenciais não são nulos ou vazios
        if descricao_classe and sigla_classe and codigo and descricao and unidade_medida and preco_desonerado is not None:
            # Adicionar a composição principal no banco de dados
            composicao_id = adicionar_composicao_principal(descricao_classe, sigla_classe, codigo, descricao,
                                                           unidade_medida, tipo, data_cotacao, preco_desonerado)
            if composicao_id:
                print(f'Composição {descricao} com código {codigo} inserida com sucesso, ID: {composicao_id}')
            else:
                print(f"Erro ao inserir composição com código {codigo}. Dados inválidos.")
        else:
            print(f"Composição com código {codigo} ignorada, pois contém dados inválidos.")


# Caminho do arquivo Excel
caminho_arquivo = "C:\\Users\\Dell\\OneDrive\\Documentos\\Projetos\\serpra\\desonerado sinapi\\SINAPI_Custo_Ref_Composicoes_Sintetico_MT_202411_Desonerado.xlsx"
# Processar o arquivo Excel
processar_composicoes_excel(caminho_arquivo)
