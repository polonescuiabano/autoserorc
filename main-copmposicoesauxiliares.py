import openpyxl
from pymongo import MongoClient
from datetime import datetime

client = MongoClient('mongodb://localhost:27017/')
db = client['serorc']
composicoes_collection = db['composicoes']
insumos_collection = db['insumos']


# Função para adicionar composição auxiliar ou insumo à composição principal
def adicionar_composicao_auxiliar_ou_insumo(composicao_principal_id, tipo, codigo_item, coeficiente):
    print(f"Processando: Tipo: {tipo}, Código Item: {codigo_item}, Coeficiente: {coeficiente}")

    if tipo == "COMPOSICAO":
        # Adicionar composição auxiliar
        if codigo_item and coeficiente is not None:
            # Garantir que o código da composição auxiliar seja inteiro
            try:
                codigo_item_int = int(codigo_item)  # Convertendo para inteiro
            except ValueError:
                print(f"Erro ao converter o código da composição auxiliar para inteiro: {codigo_item}. Ignorando.")
                return

            composicao_auxiliar = {
                "codigo": codigo_item_int,
                "coeficiente": coeficiente
            }
            print(f"Adicionando composição auxiliar: {composicao_auxiliar}")
            # Atualizar a composição principal com a composição auxiliar, se ela ainda não existir
            result = composicoes_collection.update_one(
                {"_id": composicao_principal_id, "composicoes_auxiliares.codigo": {"$ne": codigo_item_int}},
                {"$push": {"composicoes_auxiliares": composicao_auxiliar}}
            )
            if result.modified_count > 0:
                print(f"Composição auxiliar com código {codigo_item_int} foi adicionada.")
            else:
                print(f"Composição auxiliar com código {codigo_item_int} já existe ou não foi adicionada.")
        else:
            print(f"Composição auxiliar com código {codigo_item} não tem coeficiente válido (None).")

    elif tipo == "INSUMO":
        # Buscar o insumo pelo código
        if codigo_item and coeficiente is not None:
            codigo_item = codigo_item.strip()  # Remove espaços extras
            print(f"Buscando insumo com código: {codigo_item}")
            # Verificando se o código é numérico ou string
            try:
                codigo_item_int = int(codigo_item)  # Se for possível, tenta converter para inteiro
                print(f"Buscando insumo com código numérico: {codigo_item_int}")
                # Buscar insumo com código numérico
                insumo = insumos_collection.find_one({"codigo": codigo_item_int})
            except ValueError:
                # Se não for possível, tenta buscar como string
                print(f"Buscando insumo com código como string: {codigo_item}")
                insumo = insumos_collection.find_one({"codigo": codigo_item})

            if insumo:
                print(f"Insumo encontrado: {insumo}")
                insumo_com_coeficiente = {
                    "insumo_id": insumo["_id"],  # Referência ao insumo
                    "coeficiente": coeficiente
                }
                print(f"Adicionando insumo: {insumo_com_coeficiente}")
                result = composicoes_collection.update_one(
                    {"_id": composicao_principal_id, "insumos.insumo_id": {"$ne": insumo["_id"]}},
                    {"$push": {"insumos": insumo_com_coeficiente}}
                )
                if result.modified_count > 0:
                    print(f"Insumo com código {codigo_item} foi adicionado à composição principal.")
                else:
                    print(f"Insumo com código {codigo_item} já foi adicionado ou não foi adicionado.")
            else:
                print(f"Insumo com código {codigo_item} não encontrado. Tentando imprimir para depuração.")
                # Para depuração, mostre como o código está armazenado
                insumos = insumos_collection.find({"codigo": {"$regex": codigo_item, "$options": "i"}})
                for item in insumos:
                    print(f"Possíveis insumos encontrados com o código aproximado: {item}")
        else:
            print(f"Insumo com código {codigo_item} não tem coeficiente válido (None).")


# Função para processar as composições auxiliares e insumos
def processar_composicoes_auxiliares_e_insumos(caminho_arquivo):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(caminho_arquivo)
    sheet = wb.active  # Pegando a primeira aba da planilha

    # Iterar pelas linhas do Excel a partir da linha 7 até a linha 48725
    for row in range(7, 48725):  # Ajustado para começar na linha 7 e terminar na linha 48725
        codigo_composicao_principal = sheet.cell(row=row, column=7).value  # Código da composição principal (coluna G)
        tipo_item = sheet.cell(row=row, column=12).value  # Tipo (COMPOSICAO ou INSUMO) - coluna L
        codigo_item = sheet.cell(row=row, column=13).value  # Código do item (composição auxiliar ou insumo) - coluna M
        coeficiente = sheet.cell(row=row, column=17).value  # Coeficiente - coluna Q

        # Adicionar depuração para verificação dos valores processados
        print(
            f"Processando linha {row}: Código Composição Principal: {codigo_composicao_principal}, Tipo Item: {tipo_item}, Código Item: {codigo_item}, Coeficiente: {coeficiente}")

        if tipo_item in ["COMPOSICAO", "INSUMO"]:
            # Certifique-se de que o código da composição principal seja inteiro
            if isinstance(codigo_composicao_principal, str):
                codigo_composicao_principal = int(codigo_composicao_principal)

            # Buscar a composição principal no banco de dados pelo código
            composicao_principal = composicoes_collection.find_one({"codigo": codigo_composicao_principal})

            if composicao_principal:
                # Adicionar a composição auxiliar ou insumo à composição principal
                coeficiente = processar_coeficiente(coeficiente)
                if coeficiente is not None:
                    adicionar_composicao_auxiliar_ou_insumo(composicao_principal["_id"], tipo_item, codigo_item, coeficiente)
                    print(f'{tipo_item} com código {codigo_item} e coeficiente {coeficiente} foi adicionado à composição principal {codigo_composicao_principal}.')
                else:
                    print(f"Coeficiente inválido para o item {codigo_item}. Ignorando.")
            else:
                print(f'Composição principal com código {codigo_composicao_principal} não encontrada.')


# Função para processar o coeficiente
def processar_coeficiente(coeficiente):
    if coeficiente is not None:
        coef_str = str(coeficiente).strip()

        # Remover ponto como separador de milhar
        coef_str = coef_str.replace(".", "")

        # Substituir a vírgula por ponto para garantir que seja tratado como decimal
        coef_str = coef_str.replace(",", ".")

        try:
            # Tentar converter o valor para float
            return float(coef_str)
        except ValueError:
            print(f"Valor inválido para conversão para float: {coef_str}. Ignorando.")
            return None
    return None


# Caminho do arquivo Excel
caminho_arquivo = "C:\\Users\\Dell\\OneDrive\\Documentos\\Projetos\\serpra\\desonerado sinapi\\SINAPI_Custo_Ref_Composicoes_Analitico_MT_202411_Desonerado.xlsx"

# Processar as composições auxiliares e insumos
processar_composicoes_auxiliares_e_insumos(caminho_arquivo)
